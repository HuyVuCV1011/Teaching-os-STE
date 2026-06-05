import io
import csv
import openpyxl
from docx import Document
from pypdf import PdfReader
from typing import Dict, Any

def parse_pdf_to_markdown(file_bytes: bytes, title: str = "PDF Document") -> str:
    pdf_file = io.BytesIO(file_bytes)
    reader = PdfReader(pdf_file)
    md_parts = [f"# {title}\n"]
    for i, page in enumerate(reader.pages):
        page_text = page.extract_text()
        if page_text:
            md_parts.append(f"## Page {i + 1}\n\n{page_text.strip()}\n")
    return "\n".join(md_parts)

def parse_docx_to_markdown(file_bytes: bytes, title: str = "Word Document") -> str:
    docx_file = io.BytesIO(file_bytes)
    doc = Document(docx_file)
    md_parts = [f"# {title}\n"]
    
    in_list = False
    list_type = None
    
    from docx.text.paragraph import Paragraph
    from docx.table import Table
    
    def iter_block_items(parent):
        from docx.document import Document as DocumentClass
        if isinstance(parent, DocumentClass):
            parent_elm = parent.element.body
        else:
            parent_elm = parent._element
        for child in parent_elm.iterchildren():
            if child.tag.endswith('p'):
                yield Paragraph(child, parent)
            elif child.tag.endswith('tbl'):
                yield Table(child, parent)

    for block in iter_block_items(doc):
        if isinstance(block, Paragraph):
            text = block.text.strip()
            if not text:
                continue
            
            style_name = block.style.name.lower()
            is_list_item = False
            current_list_type = None
            
            if "list bullet" in style_name or text.startswith("•") or text.startswith("-"):
                is_list_item = True
                current_list_type = "ul"
                if text.startswith("•") or text.startswith("-"):
                    text = text[1:].strip()
            elif "list number" in style_name or (text and text[0].isdigit() and (text.startswith(tuple(f"{i}." for i in range(10))) or text.startswith(tuple(f"{i})" for i in range(10))))):
                is_list_item = True
                current_list_type = "ol"
                parts = text.split(".", 1)
                if len(parts) > 1 and parts[0].isdigit():
                    text = parts[1].strip()
                else:
                    parts = text.split(")", 1)
                    if len(parts) > 1 and parts[0].isdigit():
                        text = parts[1].strip()

            if is_list_item:
                if not in_list or list_type != current_list_type:
                    in_list = True
                    list_type = current_list_type
            else:
                in_list = False
                list_type = None

            md_text = ""
            for run in block.runs:
                md_run_text = run.text
                if not md_run_text:
                    continue
                is_bold = run.bold
                is_italic = run.italic
                is_underline = run.underline
                
                run_md = md_run_text
                if is_bold and is_italic:
                    run_md = f"***{run_md}***"
                elif is_bold:
                    run_md = f"**{run_md}**"
                elif is_italic:
                    run_md = f"*{run_md}*"
                if is_underline:
                    run_md = f"<u>{run_md}</u>"
                md_text += run_md

            if "heading 1" in style_name:
                md_parts.append(f"# {md_text}\n")
            elif "heading 2" in style_name:
                md_parts.append(f"## {md_text}\n")
            elif "heading 3" in style_name:
                md_parts.append(f"### {md_text}\n")
            elif "heading 4" in style_name:
                md_parts.append(f"#### {md_text}\n")
            elif is_list_item:
                bullet = "-" if list_type == "ul" else "1."
                md_parts.append(f"{bullet} {md_text}")
            else:
                md_parts.append(f"{md_text}\n")
                
        elif isinstance(block, Table):
            in_list = False
            list_type = None
            md_table_rows = []
            
            for r_idx, row in enumerate(block.rows):
                row_cells = []
                for cell in row.cells:
                    cell_text = cell.text.strip().replace("\n", " ").replace("|", "\\|")
                    row_cells.append(cell_text)
                md_table_rows.append("| " + " | ".join(row_cells) + " |")
                if r_idx == 0:
                    md_table_rows.append("| " + " | ".join(["---"] * len(row_cells)) + " |")
            md_parts.append("\n" + "\n".join(md_table_rows) + "\n")
            
    return "\n".join(md_parts)

def parse_xlsx_to_markdown(file_bytes: bytes, title: str = "Excel Sheet") -> str:
    xlsx_file = io.BytesIO(file_bytes)
    wb = openpyxl.load_workbook(xlsx_file, data_only=True)
    md_parts = [f"# {title}\n"]
    
    for name in wb.sheetnames:
        sheet = wb[name]
        md_parts.append(f"## Sheet: {name}\n")
        md_table_rows = []
        max_col = sheet.max_column or 0
        if max_col == 0:
            continue
            
        for r_idx, row in enumerate(sheet.iter_rows(values_only=True)):
            row_cells = [str(val).replace("\n", " ").replace("|", "\\|") if val is not None else "" for val in row]
            if not any(row_cells):
                continue
            md_table_rows.append("| " + " | ".join(row_cells) + " |")
            if r_idx == 0:
                md_table_rows.append("| " + " | ".join(["---"] * len(row_cells)) + " |")
                
        if md_table_rows:
            md_parts.append("\n".join(md_table_rows) + "\n")
            
    return "\n".join(md_parts)

def parse_csv_to_markdown(file_bytes: bytes, title: str = "CSV Data") -> str:
    text_content = file_bytes.decode("utf-8-sig", errors="ignore")
    reader = csv.reader(io.StringIO(text_content))
    md_parts = [f"# {title}\n"]
    md_table_rows = []
    
    for r_idx, row in enumerate(reader):
        row_cells = [val.replace("\n", " ").replace("|", "\\|") for val in row]
        if not any(row_cells):
            continue
        md_table_rows.append("| " + " | ".join(row_cells) + " |")
        if r_idx == 0:
            md_table_rows.append("| " + " | ".join(["---"] * len(row_cells)) + " |")
            
    if md_table_rows:
        md_parts.append("\n".join(md_table_rows) + "\n")
        
    return "\n".join(md_parts)
