import sys
import os
import json
import csv
import traceback
import re
from markitdown import MarkItDown

def render_html_table(table_lines):
    html_table = ['<div class="overflow-x-auto my-6"><table class="min-w-full border-collapse border border-slate-200">']
    for idx, line in enumerate(table_lines):
        # Skip divider line (e.g., | --- | --- |)
        if idx == 1 and all(c in ' |-:' for c in line.strip()):
            continue
        cells = [c.strip() for c in line.split('|')[1:-1]]
        tag = 'th' if idx == 0 else 'td'
        class_attr = ' class="border border-slate-300 px-4 py-2 bg-slate-50 text-left font-semibold text-slate-700"' if tag == 'th' else ' class="border border-slate-300 px-4 py-2 text-slate-600"'
        row_str = '<tr>' + ''.join(f'<{tag}{class_attr}>{c}</{tag}>' for c in cells) + '</tr>'
        html_table.append(row_str)
    html_table.append('</table></div>')
    return '\n'.join(html_table)

def markdown_to_html(md_text):
    # Convert headings
    html = md_text
    html = re.sub(r'^#\s+(.+)$', r'<h1>\1</h1>', html, flags=re.MULTILINE)
    html = re.sub(r'^##\s+(.+)$', r'<h2>\1</h2>', html, flags=re.MULTILINE)
    html = re.sub(r'^###\s+(.+)$', r'<h3>\1</h3>', html, flags=re.MULTILINE)
    html = re.sub(r'^####\s+(.+)$', r'<h4>\1</h4>', html, flags=re.MULTILINE)
    
    # Convert bold & italic
    html = re.sub(r'\*\*\*(.*?)\*\*\*', r'<strong><em>\1</em></strong>', html)
    html = re.sub(r'\*\*(.*?)\*\*', r'<strong>\1</strong>', html)
    html = re.sub(r'\*(.*?)\*', r'<em>\1</em>', html)
    
    # Convert tables
    lines = html.split('\n')
    in_table = False
    table_lines = []
    output_lines = []
    
    for line in lines:
        if line.strip().startswith('|'):
            if not in_table:
                in_table = True
                table_lines = [line]
            else:
                table_lines.append(line)
        else:
            if in_table:
                in_table = False
                output_lines.append(render_html_table(table_lines))
                table_lines = []
            output_lines.append(line)
    if in_table:
        output_lines.append(render_html_table(table_lines))
        
    # Standard paragraphs for text lines
    final_lines = []
    for line in output_lines:
        stripped = line.strip()
        if not stripped:
            continue
        if stripped.startswith('<') or stripped.startswith('*') or stripped.startswith('-') or stripped.startswith('1.'):
            final_lines.append(line)
        else:
            final_lines.append(f'<p>{line}</p>')
            
    return '\n'.join(final_lines)

# Keep parse_csv and parse_xlsx for tabular preview compatibility with the frontend
def parse_csv(csv_path):
    headers = []
    rows = []
    total_rows = 0
    total_cols = 0
    
    with open(csv_path, mode='r', encoding='utf-8-sig', errors='ignore') as f:
        reader = csv.reader(f)
        try:
            headers = next(reader)
            total_cols = len(headers)
        except StopIteration:
            headers = []
            
        for row in reader:
            total_rows += 1
            if len(rows) < 15:
                rows.append(row)
                
    extracted_lines = [", ".join(headers)] if headers else []
    for r in rows:
        extracted_lines.append(", ".join(r))
    extracted_text = "\n".join(extracted_lines)
    
    return {
        "headers": headers,
        "rows": rows,
        "row_count": total_rows,
        "col_count": total_cols,
        "sheet_names": []
    }, extracted_text

def parse_xlsx(xlsx_path):
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    sheet_names = wb.sheetnames
    
    sheet = wb.active if wb.active else wb[sheet_names[0]]
    
    headers = []
    rows = []
    total_rows = 0
    total_cols = sheet.max_column
    
    for r_idx, row in enumerate(sheet.iter_rows(values_only=True)):
        if r_idx == 0:
            headers = [str(val) if val is not None else "" for val in row]
        else:
            total_rows += 1
            if len(rows) < 15:
                rows.append([str(val) if val is not None else "" for val in row])
                
    extracted_lines = [", ".join(headers)] if headers else []
    for r in rows:
        extracted_lines.append(", ".join(r))
    extracted_text = "\n".join(extracted_lines)
    
    return {
        "headers": headers,
        "rows": rows,
        "row_count": total_rows,
        "col_count": total_cols,
        "sheet_names": sheet_names
    }, extracted_text

def main():
    if len(sys.argv) < 2:
        print(json.dumps({"error": "No file path provided"}), file=sys.stderr)
        sys.exit(1)
        
    file_path = sys.argv[1]
    if not os.path.exists(file_path):
        print(json.dumps({"error": f"File path does not exist: {file_path}"}), file=sys.stderr)
        sys.exit(1)
        
    _, ext = os.path.splitext(file_path.lower())
    
    try:
        if ext in ['.docx', '.pdf', '.pptx', '.zip', '.html', '.htm']:
            md = MarkItDown()
            result = md.convert(file_path)
            markdown_content = result.text_content or ""
            
            output = {
                "viewer_artifact": {
                    "type": "docx" if ext == ".docx" else "pdf" if ext == ".pdf" else "document",
                    "viewer_markdown": markdown_content,
                    "viewer_html": markdown_to_html(markdown_content)
                },
                "extracted_text": markdown_content
            }
        elif ext == '.csv':
            preview, text = parse_csv(file_path)
            output = {
                "viewer_artifact": {
                    "type": "tabular",
                    **preview
                },
                "extracted_text": text
            }
        elif ext in ['.xlsx', '.xls']:
            preview, text = parse_xlsx(file_path)
            output = {
                "viewer_artifact": {
                    "type": "tabular",
                    **preview
                },
                "extracted_text": text
            }
        else:
            print(json.dumps({"error": f"Unsupported file extension: {ext}"}), file=sys.stderr)
            sys.exit(1)
            
        print(json.dumps(output))
        
    except Exception as e:
        err_msg = f"Error processing file: {str(e)}\n{traceback.format_exc()}"
        print(json.dumps({"error": err_msg}), file=sys.stderr)
        sys.exit(1)

if __name__ == "__main__":
    main()

